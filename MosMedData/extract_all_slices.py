import numpy as np
import cv2, os, json
import nibabel as nib
from scipy import ndimage
from tqdm import tqdm
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split

def read_process_nii(data_nii_p):
    # load image and preprocess
    scan = nib.load(data_nii_p)
    volume = scan.get_fdata() # w, h, slice
    # normalize the volume
    min = -1000
    max = 400
    volume[volume < min] = min
    volume[volume > max] = max
    volume = (volume - min) / (max - min)
    volume = volume.astype("float32")
    # rotate
    volume = ndimage.rotate(volume, 90, reshape=False)
    # resize across z-axis -> 256 * 256 * 64
    assert volume.shape[:2] == (512, 512)
    width_factor = 1 / (volume.shape[0] / 256)
    height_factor = 1 / (volume.shape[1] / 256)
    volume = ndimage.zoom(volume, (width_factor, height_factor, 1.0), order=1)
    volume = (volume * 255.0).astype(np.uint8)
    return volume

if __name__ == "__main__":
    # load patient info and split the dataset into training and testing sets (8:2)
    all_data = []
    all_labels = []
    for row in load_workbook('MosMedData/dataset_registry.xlsx').active.iter_rows(min_row=2):
        all_data.append(row[2].value)
        # merge CT-234 in our work
        if row[1].value == "CT-2" or row[1].value == "CT-3" or row[1].value == "CT-4":
            all_labels.append("CT-234")
        else:
            all_labels.append(row[1].value)

    train_data, test_data, train_label, test_label = train_test_split(all_data, all_labels, test_size=0.2, stratify=all_labels, random_state=33)
    total_stat = {"training": (train_data, train_label), "testing": (test_data, test_label)}

    # At the patient level, count the number of cases in each disease category
    level, count = np.unique(train_label, return_counts=True)
    print("train cls stat:")
    for j in range(len(level)):
        print(f"{level[j]}: {count[j]}")
    print("**************")
    level, count = np.unique(test_label, return_counts=True)
    print("test cls stat:")
    for j in range(len(level)):
        print(f"{level[j]}: {count[j]}")

    # nii -> png (save all extracted slices)
    data_im = "MosMedData/MosMed_all_slices"
    os.makedirs(data_im, exist_ok=True)
    meta_file_train, meta_file_test = [], []
    
    slice_count = []
    for mode, mode_info in total_stat.items():
        with tqdm(total=len(mode_info[0]), desc=f"{mode}:") as pbar:
            for data_nii_p, label in zip(*mode_info):
                data_nii_p = "MosMedData" + data_nii_p
                assert os.path.exists(data_nii_p)
                volume = read_process_nii(data_nii_p) # w, h, slice
                assert volume.shape[:2] == (256, 256)

                slice_count.append(volume.shape[2])

                # Save the metadata of all slices, including file names and labels
                volume_name = data_nii_p.split("/")[-1].split(".")[0]
                for i in range(volume.shape[2]):
                    cv2.imwrite(f"{data_im}/{volume_name}_slice{str(i+1).zfill(2)}.png", volume[:, :, i]) 
                    
                    if mode == "training":
                        meta_file_train.append({"file_name": f"{volume_name}_slice{str(i+1).zfill(2)}.png",
                                                "cls": label
                                            })
                    elif mode == "testing":
                        meta_file_test.append({"file_name": f"{volume_name}_slice{str(i+1).zfill(2)}.png",
                                                "cls": label
                                            })
                    else:
                        raise ValueError
                
                pbar.update(1)
            pbar.close()
    
    print("train img num: ", len(meta_file_train))
    print("test img num: ", len(meta_file_test))
    print("avg slice num: ", np.mean(slice_count))
    print("min slice num: ", np.min(slice_count))
    print("max slice num: ", np.max(slice_count))

    with open(f"{data_im}/train_metadata.jsonl", 'w') as f_new:
        for item in meta_file_train:
            json.dump(item, f_new)
            f_new.write('\n')
    f_new.close()

    with open(f"{data_im}/test_metadata.jsonl", 'w') as f:
        for item in meta_file_test:
            json.dump(item, f)
            f.write('\n')
    f.close()